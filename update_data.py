# -*- coding: utf-8 -*-
import openpyxl
import json
import datetime
import re
import os
import sys

# Ensure stdout handles Thai characters properly
sys.stdout.reconfigure(encoding='utf-8')

def parse_excel():
    excel_path = r'd:\55.ลา\ระบบบันทึกข้อมูลการลาออนไลน์.xlsx'
    if not os.path.exists(excel_path):
        print(f"ไม่พบไฟล์ Excel ในตำแหน่ง: {excel_path}")
        sys.exit(1)
        
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    sheet_name = 'ข้อมูลการลา'
    if sheet_name not in wb.sheetnames:
        print(f"ไม่พบชีตชื่อ '{sheet_name}' ในไฟล์ Excel")
        sys.exit(1)
        
    ws = wb[sheet_name]
    records = []
    
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx == 0:
            continue
        if not any(row):
            continue
            
        name = row[1]
        if not name:
            continue
            
        name = re.sub(r'\s+', ' ', str(name).strip())
        position = str(row[2]).strip() if row[2] else ""
        
        try:
            fiscal_year = int(float(row[3])) if row[3] is not None else 2569
        except ValueError:
            fiscal_year = 2569
            
        try:
            round_no = int(float(row[4])) if row[4] is not None else 1
        except ValueError:
            round_no = 1
            
        round_detail = str(row[5]).strip() if row[5] else f"รอบ {round_no}"
        leave_type = str(row[6]).strip() if row[6] else ""
        
        # Helper to format dates
        def format_date(val):
            if isinstance(val, (datetime.datetime, datetime.date)):
                return val.strftime('%Y-%m-%d')
            elif val:
                val_str = str(val).strip()
                m = re.match(r'^(\d{4}-\d{2}-\d{2})', val_str)
                if m:
                    return m.group(1)
                return val_str
            return ""
            
        date_from = format_date(row[7])
        date_to = format_date(row[8])
        
        try:
            days = float(row[9]) if row[9] is not None else 0.0
            if days.is_integer():
                days = int(days)
        except ValueError:
            days = 0
            
        remark = str(row[10]).strip() if row[10] else ""
        timestamp = str(row[0]).strip() if row[0] else ""
        
        records.append({
            "timestamp": timestamp,
            "name": name,
            "position": position,
            "fiscalYear": fiscal_year,
            "round": round_no,
            "roundDetail": round_detail,
            "type": leave_type,
            "dateFrom": date_from,
            "dateTo": date_to,
            "days": days,
            "remark": remark
        })
        
    return records

if __name__ == '__main__':
    print("กำลังเริ่มอ่านข้อมูลการลาจากไฟล์ Excel...")
    records = parse_excel()
    
    # Generate Thai formatted timestamp
    thai_months_full = [
        "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
        "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"
    ]
    now = datetime.datetime.now()
    thai_year = now.year + 543
    thai_month = thai_months_full[now.month - 1]
    time_str = now.strftime('%H:%M น.')
    last_updated_str = f"{now.day} {thai_month} {thai_year} เวลา {time_str}"
    
    js_content = f"""// ข้อมูลการลาที่อัปเดตอัตโนมัติจากไฟล์ Excel
const lastUpdated = "{last_updated_str}";
const leaveRecords = {json.dumps(records, ensure_ascii=False, indent=2)};
"""
    
    output_path = r'd:\55.ลา\data.js'
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(js_content)
        
    print(f"ดึงข้อมูลสำเร็จ! พบรายการลาทั้งหมด {len(records)} รายการ")
    print(f"อัปเดตไฟล์ข้อมูลเรียบร้อย: {output_path}")
